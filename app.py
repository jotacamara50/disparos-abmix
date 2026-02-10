import os
from functools import wraps
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash

from models import Allocation, DailyReport, Vendor, db, User

try:
    from fpdf import FPDF
except Exception:
    FPDF = None

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import resend
except Exception:
    resend = None


login_manager = LoginManager()


def create_app():
    app = Flask(__name__, instance_relative_config=True)

    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-change-me")
    db_path = os.path.join(app.instance_path, "app.db")
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    os.makedirs(app.instance_path, exist_ok=True)

    db.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = "login"

    with app.app_context():
        db.create_all()

    register_routes(app)
    register_cli(app)

    return app


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def register_cli(app):
    @app.cli.command("create-user")
    def create_user_cmd():
        """Create a user: flask create-user"""
        username = input("Username: ").strip()
        password = input("Password: ").strip()
        role = input("Role (editor/viewer): ").strip().lower()
        if role not in {"editor", "viewer"}:
            print("Invalid role. Use editor or viewer.")
            return
        if User.query.filter_by(username=username).first():
            print("User already exists.")
            return
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
        )
        db.session.add(user)
        db.session.commit()
        print("User created.")

    @app.cli.command("create-vendor")
    def create_vendor_cmd():
        """Create a vendor: flask create-vendor"""
        name = input("Name: ").strip()
        email = input("Email: ").strip()
        if not name or not email:
            print("Name and email are required.")
            return
        vendor = Vendor(name=name, email=email, active=True)
        db.session.add(vendor)
        db.session.commit()
        print("Vendor created.")

    @app.cli.command("seed-vendors")
    def seed_vendors_cmd():
        """Seed initial vendors: flask seed-vendors"""
        seed_vendors = [
            {"name": "Ana Caroline Terto", "email": "comercial14@abmix.com.br"},
            {"name": "Bruno Garcia", "email": "comercial10@abmix.com.br"},
            {"name": "Fabiana Godinho", "email": "comercial9@abmix.com.br"},
            {"name": "Fernanda Batista", "email": "comercial18@abmix.com.br"},
            {"name": "Gabrielle Fernandes", "email": "comercial3@abmix.com.br"},
            {"name": "Isabela Velasquez", "email": "comercial4@abmix.com.br"},
            {"name": "Isak Rocha", "email": "comercial1@abmix.com.br"},
            {"name": "Juliana Araujo", "email": "comercial6@abmix.com.br"},
            {"name": "Kailayne Oliveira", "email": "comercial20@abmix.com.br"},
            {"name": "Luciana Velasquez", "email": "comercial21@abmix.com.br"},
            {"name": "Monique Silva", "email": "comercial2@abmix.com.br"},
            {"name": "Rodrigo Ribas", "email": "supervisao@abmix.com.br"},
            {"name": "Sara Mattos", "email": "comercial8@abmix.com.br"},
            {"name": "Michelle Manieri", "email": "michelle@abmix.com.br"},
        ]

        created = 0
        for vendor_data in seed_vendors:
            existing = Vendor.query.filter_by(email=vendor_data["email"]).first()
            if existing:
                continue
            vendor = Vendor(name=vendor_data["name"], email=vendor_data["email"], active=True)
            db.session.add(vendor)
            created += 1

        if created:
            db.session.commit()
        print(f"Seed concluido. Novos vendedores: {created}.")


def roles_required(required_role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role != required_role:
                flash("Acesso restrito.", "error")
                return redirect(url_for("dashboard"))
            return fn(*args, **kwargs)

        return wrapper

    return decorator


def parse_date(value):
    if not value or value == "None":
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        try:
            return datetime.strptime(value, "%d/%m/%Y").date()
        except ValueError:
            return None


def register_routes(app):
    @app.route("/")
    def index():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        return redirect(url_for("login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(username=username).first()
            if not user or not check_password_hash(user.password_hash, password):
                flash("Usuario ou senha invalidos.", "error")
                return render_template("login.html")
            login_user(user)
            return redirect(url_for("dashboard"))
        return render_template("login.html")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    @app.route("/dashboard")
    @login_required
    def dashboard():
        start = parse_date(request.args.get("start"))
        end = parse_date(request.args.get("end"))

        query = DailyReport.query
        if start:
            query = query.filter(DailyReport.report_date >= start)
        if end:
            query = query.filter(DailyReport.report_date <= end)

        reports = query.order_by(DailyReport.report_date.desc()).all()

        totals = {
            "total_sent": sum(r.total_sent for r in reports),
            "total_accepted": sum(r.total_accepted for r in reports),
        }

        vendor_totals = (
            db.session.query(Vendor.name, db.func.sum(Allocation.accepted_count))
            .join(Allocation, Allocation.vendor_id == Vendor.id)
            .join(DailyReport, Allocation.report_id == DailyReport.id)
        )
        if start:
            vendor_totals = vendor_totals.filter(DailyReport.report_date >= start)
        if end:
            vendor_totals = vendor_totals.filter(DailyReport.report_date <= end)
        vendor_totals = vendor_totals.group_by(Vendor.name).all()

        report_allocations = {}
        for report in reports:
            allocation_parts = []
            for allocation in sorted(report.allocations, key=lambda a: a.vendor.name):
                allocation_parts.append(f"{allocation.vendor.name}: {allocation.accepted_count}")
            report_allocations[report.id] = ", ".join(allocation_parts) if allocation_parts else "-"

        return render_template(
            "dashboard.html",
            reports=reports,
            report_allocations=report_allocations,
            totals=totals,
            vendor_totals=vendor_totals,
            start=start,
            end=end,
        )

    @app.route("/reports/new", methods=["GET", "POST"])
    @login_required
    @roles_required("editor")
    def new_report():
        vendors = Vendor.query.filter_by(active=True).order_by(Vendor.name).all()
        if request.method == "POST":
            report_date = parse_date(request.form.get("report_date"))
            total_sent = request.form.get("total_sent")
            total_accepted = request.form.get("total_accepted")
            notes = request.form.get("notes", "").strip()

            error = validate_report(report_date, total_sent, total_accepted, vendors, request.form)
            if error:
                flash(error, "error")
                return render_template("report_form.html", vendors=vendors, report=None)

            total_sent = int(total_sent)
            total_accepted = int(total_accepted)

            if DailyReport.query.filter_by(report_date=report_date).first():
                flash("Ja existe um lancamento para essa data.", "error")
                return render_template("report_form.html", vendors=vendors, report=None)

            report = DailyReport(
                report_date=report_date,
                total_sent=total_sent,
                total_accepted=total_accepted,
                notes=notes,
                created_by=current_user.id,
                updated_at=datetime.utcnow(),
            )
            db.session.add(report)
            db.session.flush()

            allocations = build_allocations(vendors, request.form, report.id)
            db.session.add_all(allocations)
            db.session.commit()

            flash("Lancamento salvo.", "success")
            return redirect(url_for("dashboard"))

        return render_template("report_form.html", vendors=vendors, report=None)

    @app.route("/reports/<int:report_id>/edit", methods=["GET", "POST"])
    @login_required
    @roles_required("editor")
    def edit_report(report_id):
        report = DailyReport.query.get_or_404(report_id)
        vendors = Vendor.query.order_by(Vendor.name).all()
        allocations_map = {a.vendor_id: a.accepted_count for a in report.allocations}

        if request.method == "POST":
            report_date = parse_date(request.form.get("report_date"))
            total_sent = request.form.get("total_sent")
            total_accepted = request.form.get("total_accepted")
            notes = request.form.get("notes", "").strip()

            error = validate_report(report_date, total_sent, total_accepted, vendors, request.form)
            if error:
                flash(error, "error")
                return render_template(
                    "report_form.html",
                    vendors=vendors,
                    report=report,
                    allocations_map=allocations_map,
                )

            total_sent = int(total_sent)
            total_accepted = int(total_accepted)

            existing = DailyReport.query.filter_by(report_date=report_date).first()
            if existing and existing.id != report.id:
                flash("Ja existe um lancamento para essa data.", "error")
                return render_template(
                    "report_form.html",
                    vendors=vendors,
                    report=report,
                    allocations_map=allocations_map,
                )

            report.report_date = report_date
            report.total_sent = total_sent
            report.total_accepted = total_accepted
            report.notes = notes
            report.updated_at = datetime.utcnow()

            Allocation.query.filter_by(report_id=report.id).delete()
            allocations = build_allocations(vendors, request.form, report.id)
            db.session.add_all(allocations)
            db.session.commit()

            flash("Lancamento atualizado.", "success")
            return redirect(url_for("dashboard"))

        return render_template(
            "report_form.html",
            vendors=vendors,
            report=report,
            allocations_map=allocations_map,
        )

    @app.route("/vendors", methods=["GET", "POST"])
    @login_required
    @roles_required("editor")
    def vendors():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            email = request.form.get("email", "").strip()
            if not name or not email:
                flash("Nome e email sao obrigatorios.", "error")
            else:
                vendor = Vendor(name=name, email=email, active=True)
                db.session.add(vendor)
                db.session.commit()
                flash("Vendedor criado.", "success")
            return redirect(url_for("vendors"))

        vendor_list = Vendor.query.order_by(Vendor.name).all()
        return render_template("vendors.html", vendors=vendor_list)

    @app.route("/vendors/<int:vendor_id>/edit", methods=["GET", "POST"])
    @login_required
    @roles_required("editor")
    def edit_vendor(vendor_id):
        vendor = Vendor.query.get_or_404(vendor_id)
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            email = request.form.get("email", "").strip()
            active = request.form.get("active") == "on"
            if not name or not email:
                flash("Nome e email sao obrigatorios.", "error")
                return render_template("vendor_edit.html", vendor=vendor)
            vendor.name = name
            vendor.email = email
            vendor.active = active
            db.session.commit()
            flash("Vendedor atualizado.", "success")
            return redirect(url_for("vendors"))
        return render_template("vendor_edit.html", vendor=vendor)

    @app.route("/reports/<int:report_id>/send-email", methods=["POST"])
    @login_required
    @roles_required("editor")
    def send_report_email_route(report_id):
        report = DailyReport.query.get_or_404(report_id)
        allocations = [
            {"vendor_id": allocation.vendor_id, "accepted_count": allocation.accepted_count}
            for allocation in report.allocations
        ]
        success, error = send_report_email(report, allocations)
        if success:
            flash("Email enviado.", "success")
        else:
            flash(error or "Falha ao enviar email.", "error")
        return redirect(
            url_for(
                "dashboard",
                start=request.args.get("start"),
                end=request.args.get("end"),
            )
        )

    @app.route("/export/excel")
    @login_required
    def export_excel():
        if openpyxl is None:
            flash("Exportacao Excel indisponivel (dependencia faltando).", "error")
            return redirect(url_for("dashboard"))

        start = parse_date(request.args.get("start"))
        end = parse_date(request.args.get("end"))

        query = DailyReport.query
        if start:
            query = query.filter(DailyReport.report_date >= start)
        if end:
            query = query.filter(DailyReport.report_date <= end)

        reports = query.order_by(DailyReport.report_date.asc()).all()
        vendors = Vendor.query.order_by(Vendor.name).all()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio"

        headers = ["Data", "Disparos", "Aceites"] + [v.name for v in vendors]
        sheet.append(headers)

        for report in reports:
            allocations = {a.vendor_id: a.accepted_count for a in report.allocations}
            row = [
                report.report_date.isoformat(),
                report.total_sent,
                report.total_accepted,
            ]
            row.extend([allocations.get(v.id, 0) for v in vendors])
            sheet.append(row)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = "relatorio.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/export/pdf")
    @login_required
    def export_pdf():
        if FPDF is None:
            flash("Exportacao PDF indisponivel (dependencia faltando).", "error")
            return redirect(url_for("dashboard"))

        start = parse_date(request.args.get("start"))
        end = parse_date(request.args.get("end"))

        query = DailyReport.query
        if start:
            query = query.filter(DailyReport.report_date >= start)
        if end:
            query = query.filter(DailyReport.report_date <= end)

        reports = query.order_by(DailyReport.report_date.asc()).all()

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, safe_text("Relatorio Diario"), ln=True)

        pdf.set_font("Helvetica", size=10)
        if start or end:
            period = f"Periodo: {start.isoformat() if start else 'inicio'} a {end.isoformat() if end else 'hoje'}"
            pdf.cell(0, 7, safe_text(period), ln=True)
        pdf.cell(0, 7, safe_text(f"Gerado em: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"), ln=True)
        pdf.ln(2)

        total_sent_sum = sum(r.total_sent for r in reports)
        total_accepted_sum = sum(r.total_accepted for r in reports)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, safe_text("Resumo do periodo"), ln=True)
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, safe_text(f"Disparos: {total_sent_sum} | Aceites: {total_accepted_sum}"), ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, safe_text("Relatorios diarios"), ln=True)
        pdf.set_font("Helvetica", size=10)

        body_width = pdf.w - pdf.l_margin - pdf.r_margin
        name_col = int(body_width * 0.72)
        count_col = int(body_width - name_col)

        for report in reports:
            line = (
                f"Data: {report.report_date.strftime('%d/%m/%Y')} | Disparos: {report.total_sent} | "
                f"Aceites: {report.total_accepted}"
            )
            pdf.multi_cell(body_width, 6, safe_text(line))

            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(body_width, 6, safe_text("Encaminhados"), ln=True)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(name_col, 6, safe_text("Vendedor"), border=1)
            pdf.cell(count_col, 6, safe_text("Aceites"), border=1, ln=True, align="C")

            pdf.set_font("Helvetica", size=9)
            allocations_sorted = sorted(report.allocations, key=lambda a: a.vendor.name)
            if allocations_sorted:
                for allocation in allocations_sorted:
                    name_text = truncate_text(allocation.vendor.name, 48)
                    pdf.cell(name_col, 6, safe_text(name_text), border=1)
                    pdf.cell(count_col, 6, safe_text(str(allocation.accepted_count)), border=1, ln=True, align="C")
            else:
                pdf.cell(name_col, 6, safe_text("-"), border=1)
                pdf.cell(count_col, 6, safe_text("0"), border=1, ln=True, align="C")

            if report.notes:
                pdf.ln(1)
                pdf.set_font("Helvetica", "B", 10)
                pdf.cell(body_width, 6, safe_text("Observacoes"), ln=True)
                pdf.set_font("Helvetica", size=9)
                pdf.multi_cell(body_width, 5, safe_text(report.notes))

            pdf.set_font("Helvetica", size=10)
            pdf.ln(3)

        pdf_output = pdf.output(dest="S")
        if isinstance(pdf_output, str):
            pdf_bytes = pdf_output.encode("latin-1")
        else:
            pdf_bytes = bytes(pdf_output)
        output = BytesIO(pdf_bytes)

        filename = "relatorio.pdf"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")

    @app.route("/api/n8n/report", methods=["POST"])
    def api_n8n_report():
        token = extract_api_token(request)
        api_token = os.getenv("API_TOKEN")
        if not api_token or token != api_token:
            return jsonify({"error": "unauthorized"}), 401

        payload = request.get_json(silent=True) or {}
        report_date = parse_date(payload.get("date")) or date.today()
        total_sent = payload.get("total_sent")
        total_accepted = payload.get("total_accepted")
        if total_accepted is None:
            total_accepted = payload.get("total_aceites")
        total_recusados = payload.get("total_recusados")
        notes = (payload.get("notes") or "").strip()
        allocations_payload = payload.get("allocations") or []

        try:
            total_sent = int(total_sent)
        except (TypeError, ValueError):
            return jsonify({"error": "total_sent must be an integer"}), 400

        if total_sent < 0:
            return jsonify({"error": "total_sent must be >= 0"}), 400

        allocations, allocation_sum, error = parse_allocations(allocations_payload)
        if error:
            return jsonify({"error": error}), 400

        if total_accepted is None:
            total_accepted = allocation_sum
        try:
            total_accepted = int(total_accepted)
        except (TypeError, ValueError):
            return jsonify({"error": "total_accepted must be an integer"}), 400

        if total_accepted < 0:
            return jsonify({"error": "total_accepted must be >= 0"}), 400
        if total_accepted > total_sent:
            return jsonify({"error": "total_accepted cannot exceed total_sent"}), 400
        if allocation_sum != total_accepted:
            return jsonify({"error": "allocation sum must equal total_accepted"}), 400

        api_user = resolve_api_user()
        if not api_user:
            return (
                jsonify(
                    {
                        "error": "API user not found. Create an editor user and set API_USER.",
                    }
                ),
                400,
            )

        report = DailyReport.query.filter_by(report_date=report_date).first()
        updated = False
        if report:
            report.total_sent = total_sent
            report.total_accepted = total_accepted
            report.notes = notes
            report.updated_at = datetime.utcnow()
            Allocation.query.filter_by(report_id=report.id).delete()
            updated = True
        else:
            report = DailyReport(
                report_date=report_date,
                total_sent=total_sent,
                total_accepted=total_accepted,
                notes=notes,
                created_by=api_user.id,
                updated_at=datetime.utcnow(),
            )
            db.session.add(report)
            db.session.flush()

        db.session.add_all(
            [
                Allocation(
                    report_id=report.id,
                    vendor_id=allocation["vendor_id"],
                    accepted_count=allocation["accepted_count"],
                )
                for allocation in allocations
            ]
        )
        db.session.commit()

        if not updated:
            send_report_email(report, allocations)

        response = {"status": "ok", "report_id": report.id, "updated": updated}
        if total_recusados is not None:
            response["total_recusados_ignored"] = total_recusados
        return jsonify(response), 200


def validate_report(report_date, total_sent, total_accepted, vendors, form_data):
    if not report_date:
        return "Data invalida."
    if total_sent is None or total_accepted is None:
        return "Preencha disparos e aceites."
    if not str(total_sent).isdigit() or not str(total_accepted).isdigit():
        return "Disparos e aceites devem ser numeros inteiros."

    total_sent = int(total_sent)
    total_accepted = int(total_accepted)

    if total_sent < 0 or total_accepted < 0:
        return "Valores nao podem ser negativos."
    if total_accepted > total_sent:
        return "Aceites nao podem ser maiores que disparos."

    allocation_sum = 0
    for vendor in vendors:
        value = form_data.get(f"vendor_{vendor.id}", "0").strip()
        if value == "":
            value = "0"
        if not value.isdigit():
            return "Encaminhamentos devem ser numeros inteiros."
        allocation_sum += int(value)

    if allocation_sum != total_accepted:
        return "A soma dos encaminhamentos deve ser igual aos aceites."

    return None


def build_allocations(vendors, form_data, report_id):
    allocations = []
    for vendor in vendors:
        value = form_data.get(f"vendor_{vendor.id}", "0").strip()
        if value == "":
            value = "0"
        count = int(value)
        if count > 0:
            allocations.append(
                Allocation(
                    report_id=report_id,
                    vendor_id=vendor.id,
                    accepted_count=count,
                )
            )
    return allocations


def safe_text(value):
    text = value if value is not None else ""
    return str(text).encode("latin-1", "replace").decode("latin-1")


def truncate_text(value, max_len):
    text = value if value is not None else ""
    text = str(text).strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def send_report_email(report, allocations):
    if resend is None:
        return False, "Envio de email indisponivel."

    api_key = os.getenv("RESEND_API_KEY")
    from_addr = os.getenv("RESEND_FROM")
    to_list = os.getenv("RESEND_TO", "")

    if not api_key or not from_addr or not to_list:
        return False, "Email nao configurado."

    recipients = [email.strip() for email in to_list.split(",") if email.strip()]
    if not recipients:
        return False, "Lista de emails vazia."

    resend.api_key = api_key

    report_date = report.report_date.strftime("%d/%m/%Y")
    subject = f"Relatorio Diario de Disparos - {report_date}"

    allocation_rows = ""
    allocations_sorted = sorted(allocations, key=lambda a: a["vendor_id"])
    if allocations_sorted:
        rows = []
        for allocation in allocations_sorted:
            vendor = Vendor.query.get(allocation["vendor_id"])
            if not vendor:
                continue
            rows.append(
                f"<tr>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;'>{vendor.name}</td>"
                f"<td style='padding:8px;border:1px solid #e5e7eb;text-align:center;'>{allocation['accepted_count']}</td>"
                f"</tr>"
            )
        allocation_rows = "".join(rows)
    else:
        allocation_rows = (
            "<tr>"
            "<td style='padding:8px;border:1px solid #e5e7eb;'>-</td>"
            "<td style='padding:8px;border:1px solid #e5e7eb;text-align:center;'>0</td>"
            "</tr>"
        )

    html = f"""
<!doctype html>
<html>
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Relatorio Diario</title>
  </head>
  <body style="margin:0;padding:0;background:#f5f7fb;font-family:Arial,Helvetica,sans-serif;color:#111827;">
    <div style="max-width:720px;margin:0 auto;padding:24px;">
      <div style="background:#0b6e4f;color:#ffffff;padding:20px 24px;border-radius:12px 12px 0 0;">
        <h1 style="margin:0;font-size:20px;font-weight:700;">Relatorio Diario de Disparos</h1>
        <p style="margin:6px 0 0;font-size:14px;">Data: {report_date}</p>
      </div>
      <div style="background:#ffffff;padding:20px 24px;border-radius:0 0 12px 12px;box-shadow:0 8px 24px rgba(15,23,42,0.08);">
        <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;">
          <div style="flex:1;min-width:200px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;">
            <div style="font-size:12px;text-transform:uppercase;color:#6b7280;letter-spacing:1px;">Disparos</div>
            <div style="font-size:26px;font-weight:700;margin-top:6px;">{report.total_sent}</div>
          </div>
          <div style="flex:1;min-width:200px;border:1px solid #e5e7eb;border-radius:10px;padding:12px;background:#f0fdfa;">
            <div style="font-size:12px;text-transform:uppercase;color:#0f766e;letter-spacing:1px;">Aceites</div>
            <div style="font-size:26px;font-weight:700;margin-top:6px;color:#0f766e;">{report.total_accepted}</div>
          </div>
        </div>
        <h2 style="font-size:16px;margin:0 0 8px;">Encaminhamentos por vendedor</h2>
        <table style="width:100%;border-collapse:collapse;font-size:14px;">
          <thead>
            <tr>
              <th style="text-align:left;padding:8px;border:1px solid #e5e7eb;background:#f8fafc;">Vendedor</th>
              <th style="text-align:center;padding:8px;border:1px solid #e5e7eb;background:#f8fafc;">Aceites</th>
            </tr>
          </thead>
          <tbody>
            {allocation_rows}
          </tbody>
        </table>
        <div style="margin-top:16px;font-size:12px;color:#6b7280;">Gerado automaticamente pelo n8n.</div>
      </div>
    </div>
  </body>
</html>
"""

    try:
        resend.Emails.send(
            {
                "from": from_addr,
                "to": recipients,
                "subject": subject,
                "html": html,
            }
        )
    except Exception:
        return False, "Falha ao enviar email."

    return True, None


def extract_api_token(req):
    auth_header = req.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return req.headers.get("X-API-Key")


def resolve_api_user():
    api_username = os.getenv("API_USER")
    if api_username:
        return User.query.filter_by(username=api_username).first()
    return User.query.filter_by(role="editor").order_by(User.id.asc()).first()


def parse_allocations(items):
    allocations = []
    allocation_sum = 0
    if not isinstance(items, list):
        return [], 0, "allocations must be a list"

    for item in items:
        if not isinstance(item, dict):
            return [], 0, "allocation items must be objects"
        vendor_email = (item.get("email") or "").strip().lower()
        vendor_name = (item.get("name") or "").strip()
        count = item.get("count")
        if count is None:
            count = item.get("leads_interessados")

        try:
            count = int(count)
        except (TypeError, ValueError):
            return [], 0, "allocation count must be an integer"

        if count < 0:
            return [], 0, "allocation count must be >= 0"

        vendor = None
        if vendor_email:
            vendor = Vendor.query.filter(db.func.lower(Vendor.email) == vendor_email).first()
        if vendor is None and vendor_name:
            vendor = Vendor.query.filter_by(name=vendor_name).first()

        if vendor is None:
            return [], 0, f"vendor not found: {vendor_email or vendor_name}"

        if count > 0:
            allocations.append({"vendor_id": vendor.id, "accepted_count": count})
            allocation_sum += count

    return allocations, allocation_sum, None


app = create_app()


if __name__ == "__main__":
    app.run(debug=True)
